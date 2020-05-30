####################################################################
# @file    : user_rating.py
# @brief   : Get the rating from the user, after the successful
#            login into the account
# @author  : Ganesh Kumar
# @date    : 8 AUG 2019
####################################################################

# Python Global Imports
import os
import sys
import openpyxl

# Python Local Imports
import database.user_details_database as user_details_database
import constants.constants as constants


class UserRating(object):
    userRating = 0
    userRatingStatus = False

    def __init__(self, user_name, global_total_number_of_booth):
        # user name can be email id or emp id
        # database should map the rating with respect to both of any user
        self.user_name = str(user_name)
        self.global_total_number_of_booth = global_total_number_of_booth

    def get_user_rating(self):
        rating_flag = constants.RATING_FLAG
        booth_number_to_vote_validity_flag = constants.BOOTH_NUMBER_FLAG
        booth_number_to_vote = 0
        while booth_number_to_vote_validity_flag is True:
            try:
                booth_number_to_vote = int(input('Please Enter the booth number'
                                                 ' to Vote: '))
                print type(booth_number_to_vote)
                print booth_number_to_vote
                print self.global_total_number_of_booth
                if (booth_number_to_vote > 0) and \
                        booth_number_to_vote < self.\
                        global_total_number_of_booth:
                    booth_number_to_vote_validity_flag = False
            except Exception:
                continue
        while rating_flag is True:
            try:
                user_rating_value = int(input('Please Enter your Rating'
                                              ' between 1 to 10: '))
                self.user_rating_excel_update(self.user_name,
                                              booth_number_to_vote,
                                              user_rating_value)
                if (user_rating_value > 0) and (user_rating_value < 11):
                    self.userRating = user_rating_value
                    self.userRatingStatus = True
                    rating_flag = False
            except Exception:
                continue

    # # Idea of implementing descriptors
    # def getUserName(self):
    #     pass
    #
    # # Idea of implementing descriptors
    # def getRatingFromUser(self):
    #     pass
    #
    # # Idea of implementing descriptors
    # def getBoothNameFromUserToRate(self):
    #     pass

    def _print_string(self):
        print str(self.user_name) + ' rating is ' + str(self.userRating)

    def user_rating_excel_update(self, user_name, booth_number, user_rating):
        try:
            user_rating_excel_file_path = user_details_database.\
                EXCEL_SHEET_FILE_PATH
            wb_obj = openpyxl.load_workbook(user_rating_excel_file_path)
            sheet_obj = wb_obj.active
            max_row = sheet_obj.max_row
            max_column = sheet_obj.max_column

            self.data_population_for_rating(user_name, booth_number,
                                            user_rating, wb_obj, sheet_obj,
                                            max_row, max_column)
            wb_obj.save(user_rating_excel_file_path)

        except Exception as e_obj:
            print 'EXCEPTION: In user_rating_excel_update(), due to %s' % e_obj

    @staticmethod
    def data_population_for_rating(user_name, booth_number, user_rating,
                                   wb_obj, sheet_obj, max_row, max_column):
        try:
            for value in range(1, max_column + 1):
                if value == 1:
                    sheet_obj.cell(row=max_row + 1, column=value).value =\
                        user_name
                elif value == int(booth_number) + 1:
                    sheet_obj.cell(row=max_row + 1, column=value).value =\
                        user_rating
                else:
                    sheet_obj.cell(row=max_row + 1, column=value).value = 0
        except Exception as e_obj:
            print 'EXCEPTION: In data_population_for_rating(), due to %s' \
                  % e_obj


