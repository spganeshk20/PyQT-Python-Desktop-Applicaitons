#######################################################################
# @file    : configure_excel_sheet.py
# @brief   : Creates the excel sheet if does not exists and configures
#            the excel sheet headers with number of booths dynamically
# @author  : Ganesh Kumar
# @date    : 16 AUG 2019
#######################################################################

# Python Global Imports
import os
from openpyxl import Workbook

# Python Local Imports
import database.user_details_database as user_details_database


def create_user_rating_excel_sheet(number_of_booths):
    try:
        user_rating_excel_file_path = user_details_database.\
            EXCEL_SHEET_FILE_PATH
        wb_obj = Workbook()
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        for data_value in range(1, number_of_booths + 2):
            if data_value == 1:
                sheet_obj.cell(row=max_row, column=data_value).value =\
                    'User Name'
            else:
                sheet_obj.cell(row=max_row, column=data_value).value =\
                    'Booth Number ' + str(data_value-1)
        wb_obj.save(user_rating_excel_file_path)
        return True

    except Exception as e_obj:
        print 'EXCEPTION: In create_user_rating_excel_sheet(), due to %s'\
              % e_obj
        return False
